"""
report_generator.py
商机扫描汇总周报生成器
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional


class WeeklyReportGenerator:
    """商机扫描汇总周报生成器"""

    # 样式定义
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=14, color="366092")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self):
        self.data: List[Dict] = []
        self.summary: Dict[str, Any] = {}
        self.high_value_opportunities: List[Dict] = []

    # ========================
    # 数据加载
    # ========================
    def load_from_excel(self, excel_path: str) -> int:
        """从验证结果 Excel 加载数据"""
        wb = openpyxl.load_workbook(excel_path)
        # 读取最新的结果 sheet
        sheet_names = [s for s in wb.sheetnames if s.startswith("验证结果_")]
        if not sheet_names:
            # 尝试读取 active sheet
            ws = wb.active
        else:
            # 读取最新的 sheet
            ws = wb[sheet_names[-1]]

        headers = [cell.value for cell in ws[1]]

        # 找到关键列
        col_map = {}
        for idx, h in enumerate(headers):
            if h:
                h_lower = str(h).lower()
                if "标题" in h:
                    col_map["标题"] = idx
                elif "最终判断" in h or "是否属于商机" in h:
                    col_map["是否属于商机"] = idx
                elif "最终推荐" in h or "推荐跟进" in h:
                    col_map["推荐跟进"] = idx
                elif "置信度" in h and "最终" in h:
                    col_map["置信度"] = idx
                elif "人工复核" in h or "人工审核" in h:
                    col_map["人工复核"] = idx
                elif "关键词" in h:
                    col_map["关键词"] = idx
                elif "商机类型" in h:
                    col_map["商机类型"] = idx
                elif "跟进优先级" in h and "最终" not in h:
                    col_map["跟进优先级"] = idx

        # 读取数据行
        self.data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = {}
            for key, idx in col_map.items():
                if idx < len(row):
                    item[key] = row[idx]
            self.data.append(item)

        return len(self.data)

    def load_from_jsonl(self, jsonl_path: str) -> int:
        """从扫描结果 JSONL 加载数据"""
        self.data = []
        with open(jsonl_path, "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    record = json.loads(line)
                    # 统一字段名
                    item = {}
                    cls = record.get("分类结果", {})

                    # 原始数据
                    orig = record.get("原始数据", {})
                    item["标题"] = orig.get("标题", "")

                    # 分类结果
                    item["是否属于商机"] = cls.get("是否属于商机", "")
                    item["推荐跟进"] = cls.get("推荐跟进", "")
                    item["跟进优先级"] = cls.get("跟进优先级", "")
                    item["置信度"] = cls.get("置信度", "")
                    item["商机类型"] = cls.get("商机类型", [])
                    item["关键词"] = cls.get("关键词", [])
                    item["判断依据"] = cls.get("判断依据", "")
                    item["风险提示"] = cls.get("风险提示", "")
                    item["人工复核"] = cls.get("是否进入人工复核池", "")

                    self.data.append(item)
                except json.JSONDecodeError:
                    continue

        return len(self.data)

    def load_from_json(self, json_path: str) -> int:
        """从扫描结果 JSON 加载数据"""
        with open(json_path, "r", encoding="utf-8") as f:
            records = json.load(f)

        if isinstance(records, list):
            self.data = records
        elif isinstance(records, dict):
            if "data" in records:
                self.data = records["data"]
            elif "results" in records:
                self.data = records["results"]
            else:
                self.data = [records]
        return len(self.data)

    def merge_data(self, other_data: List[Dict]):
        """合并外部数据"""
        self.data.extend(other_data)

    # ========================
    # 统计分析
    # ========================
    def aggregate(self) -> Dict[str, Any]:
        """汇总统计"""
        total = len(self.data)

        # 初始化统计结构
        self.summary = {
            "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "总处理数量": total,
            "商机识别结果": {"是": 0, "否": 0, "不确定": 0},
            "推荐跟进结果": {"建议跟进": 0, "可观察": 0, "不建议跟进": 0},
            "优先级分布": {"高": 0, "中": 0, "低": 0},
            "置信度分布": {"高": 0, "中": 0, "低": 0},
            "进入人工复核池": 0,
            "商机类型统计": {},
            "关键词统计": {},
        }

        for item in self.data:
            # 商机识别结果
            biz_result = item.get("是否属于商机", "")
            if biz_result in self.summary["商机识别结果"]:
                self.summary["商机识别结果"][biz_result] += 1

            # 推荐跟进
            rec = item.get("推荐跟进", "")
            if rec in self.summary["推荐跟进结果"]:
                self.summary["推荐跟进结果"][rec] += 1

            # 优先级
            priority = item.get("跟进优先级", "")
            if priority in self.summary["优先级分布"]:
                self.summary["优先级分布"][priority] += 1

            # 置信度
            conf = item.get("置信度", "")
            if conf in self.summary["置信度分布"]:
                self.summary["置信度分布"][conf] += 1

            # 人工复核
            if item.get("人工复核", "") == "是":
                self.summary["进入人工复核池"] += 1

            # 商机类型统计
            biz_types = item.get("商机类型", [])
            if isinstance(biz_types, list):
                for bt in biz_types:
                    if bt and bt != "无" and bt != "其他":
                        self.summary["商机类型统计"][bt] = self.summary["商机类型统计"].get(bt, 0) + 1

            # 关键词统计
            keywords = item.get("关键词", [])
            if isinstance(keywords, list):
                for kw in keywords:
                    if kw:
                        self.summary["关键词统计"][kw] = self.summary["关键词统计"].get(kw, 0) + 1

        # 排序关键词统计
        self.summary["商机类型统计"] = dict(
            sorted(self.summary["商机类型统计"].items(), key=lambda x: x[1], reverse=True)[:10]
        )
        self.summary["关键词统计"] = dict(
            sorted(self.summary["关键词统计"].items(), key=lambda x: x[1], reverse=True)[:20]
        )

        # 提取高价值商机
        self.high_value_opportunities = self.get_high_value_opportunities()

        return self.summary

    def get_high_value_opportunities(self) -> List[Dict]:
        """提取高价值商机（建议跟进的）"""
        high_value = []
        for item in self.data:
            if item.get("推荐跟进") == "建议跟进":
                high_value.append({
                    "标题": item.get("标题", ""),
                    "商机类型": ", ".join(item.get("商机类型", [])) if isinstance(item.get("商机类型"), list) else item.get("商机类型", ""),
                    "跟进优先级": item.get("跟进优先级", ""),
                    "置信度": item.get("置信度", ""),
                    "关键词": ", ".join(item.get("关键词", [])[:3]) if isinstance(item.get("关键词"), list) else item.get("关键词", ""),
                    "判断依据": item.get("判断依据", ""),
                    "风险提示": item.get("风险提示", ""),
                })

        # 按优先级排序
        priority_order = {"高": 0, "中": 1, "低": 2}
        high_value.sort(key=lambda x: priority_order.get(x.get("跟进优先级", ""), 3))

        return high_value

    # ========================
    # Excel 报告生成
    # ========================
    def generate_excel(self, output_path: str):
        """生成 Excel 周报"""
        wb = openpyxl.Workbook()

        # Sheet1: 统计摘要
        self._create_summary_sheet(wb)

        # Sheet2: 高价值商机
        self._create_opportunities_sheet(wb)

        # Sheet3: 详细数据
        self._create_detail_sheet(wb)

        # 保存
        wb.save(output_path)
        print(f"周报已生成: {output_path}")

    def _create_summary_sheet(self, wb):
        """创建统计摘要 Sheet"""
        ws = wb.active
        ws.title = "统计摘要"

        row = 1
        # 标题
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "商机扫描汇总周报"
        ws[f"A{row}"].font = self.TITLE_FONT
        ws[f"A{row}"].alignment = self.CENTER_ALIGN
        ws.row_dimensions[row].height = 30

        row += 1
        ws[f"A{row}"] = f"生成时间: {self.summary.get('生成时间', '')}"
        ws[f"A{row}"].font = Font(size=10, italic=True, color="666666")
        row += 2

        # 总处理数量
        ws[f"A{row}"] = "总处理数量"
        ws[f"A{row}"].fill = self.SUBHEADER_FILL
        ws[f"A{row}"].font = self.SUBHEADER_FONT
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"C{row}"] = self.summary.get("总处理数量", 0)
        ws[f"C{row}"].font = Font(bold=True, size=12)
        ws[f"C{row}"].alignment = self.CENTER_ALIGN
        row += 2

        # 商机识别结果
        ws[f"A{row}"] = "商机识别结果"
        ws[f"A{row}"].fill = self.HEADER_FILL
        ws[f"A{row}"].font = self.HEADER_FONT
        ws.merge_cells(f"A{row}:C{row}")
        for col, val in [("A", "是"), ("B", "否"), ("C", "不确定")]:
            ws[f"{col}{row}"].fill = self.HEADER_FILL
            ws[f"{col}{row}"].font = self.HEADER_FONT
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
        row += 1
        biz_results = self.summary.get("商机识别结果", {})
        ws[f"A{row}"] = biz_results.get("是", 0)
        ws[f"B{row}"] = biz_results.get("否", 0)
        ws[f"C{row}"] = biz_results.get("不确定", 0)
        for col in "ABC":
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 2

        # 推荐跟进结果
        ws[f"A{row}"] = "推荐跟进结果"
        ws[f"A{row}"].fill = self.SUBHEADER_FILL
        ws[f"A{row}"].font = self.SUBHEADER_FONT
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        rec_results = self.summary.get("推荐跟进结果", {})
        ws[f"A{row}"] = "建议跟进"
        ws[f"B{row}"] = "可观察"
        ws[f"C{row}"] = "不建议跟进"
        for col in "ABC":
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 1
        ws[f"A{row}"] = rec_results.get("建议跟进", 0)
        ws[f"B{row}"] = rec_results.get("可观察", 0)
        ws[f"C{row}"] = rec_results.get("不建议跟进", 0)
        for col in "ABC":
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 2

        # 优先级分布
        ws[f"A{row}"] = "优先级分布"
        ws[f"A{row}"].fill = self.SUBHEADER_FILL
        ws[f"A{row}"].font = self.SUBHEADER_FONT
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        priority_dist = self.summary.get("优先级分布", {})
        ws[f"A{row}"] = "高"
        ws[f"B{row}"] = "中"
        ws[f"C{row}"] = "低"
        for col in "ABC":
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 1
        ws[f"A{row}"] = priority_dist.get("高", 0)
        ws[f"B{row}"] = priority_dist.get("中", 0)
        ws[f"C{row}"] = priority_dist.get("低", 0)
        for col in "ABC":
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 2

        # 置信度分布
        ws[f"A{row}"] = "置信度分布"
        ws[f"A{row}"].fill = self.SUBHEADER_FILL
        ws[f"A{row}"].font = self.SUBHEADER_FONT
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        conf_dist = self.summary.get("置信度分布", {})
        ws[f"A{row}"] = "高"
        ws[f"B{row}"] = "中"
        ws[f"C{row}"] = "低"
        for col in "ABC":
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 1
        ws[f"A{row}"] = conf_dist.get("高", 0)
        ws[f"B{row}"] = conf_dist.get("中", 0)
        ws[f"C{row}"] = conf_dist.get("低", 0)
        for col in "ABC":
            ws[f"{col}{row}"].alignment = self.CENTER_ALIGN
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 2

        # 人工复核
        ws[f"A{row}"] = "进入人工复核池"
        ws[f"A{row}"].fill = self.SUBHEADER_FILL
        ws[f"A{row}"].font = self.SUBHEADER_FONT
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"C{row}"] = self.summary.get("进入人工复核池", 0)
        ws[f"C{row}"].alignment = self.CENTER_ALIGN
        ws[f"C{row}"].border = self.THIN_BORDER
        row += 2

        # 商机类型 Top10
        if self.summary.get("商机类型统计"):
            ws[f"A{row}"] = "商机类型统计（Top10）"
            ws[f"A{row}"].fill = self.SUBHEADER_FILL
            ws[f"A{row}"].font = self.SUBHEADER_FONT
            ws.merge_cells(f"A{row}:C{row}")
            row += 1
            for biz_type, count in list(self.summary["商机类型统计"].items())[:10]:
                ws[f"A{row}"] = biz_type
                ws[f"B{row}"] = count
                ws[f"A{row}"].border = self.THIN_BORDER
                ws[f"B{row}"].border = self.THIN_BORDER
                ws[f"B{row}"].alignment = self.CENTER_ALIGN
                row += 1

        # 调整列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _create_opportunities_sheet(self, wb):
        """创建高价值商机 Sheet"""
        ws = wb.create_sheet("高价值商机")

        # 表头
        headers = ["序号", "标题", "商机类型", "优先级", "置信度", "关键词", "判断依据", "风险提示"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER

        # 数据
        for row_idx, opp in enumerate(self.high_value_opportunities, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=2, value=opp.get("标题", "")).alignment = self.LEFT_ALIGN
            ws.cell(row=row_idx, column=3, value=opp.get("商机类型", "")).alignment = self.LEFT_ALIGN
            ws.cell(row=row_idx, column=4, value=opp.get("跟进优先级", "")).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=5, value=opp.get("置信度", "")).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=6, value=opp.get("关键词", "")).alignment = self.LEFT_ALIGN
            ws.cell(row=row_idx, column=7, value=opp.get("判断依据", "")).alignment = self.LEFT_ALIGN
            ws.cell(row=row_idx, column=8, value=opp.get("风险提示", "")).alignment = self.LEFT_ALIGN

            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = self.THIN_BORDER

        # 调整列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 40
        ws.column_dimensions["H"].width = 30

    def _create_detail_sheet(self, wb):
        """创建详细数据 Sheet"""
        ws = wb.create_sheet("详细数据")

        # 表头
        headers = ["标题", "是否属于商机", "推荐跟进", "跟进优先级", "置信度", "人工复核", "商机类型", "关键词"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER

        # 数据
        for row_idx, item in enumerate(self.data, 2):
            biz_types = item.get("商机类型", [])
            if isinstance(biz_types, list):
                biz_types_str = ", ".join(biz_types)
            else:
                biz_types_str = str(biz_types)

            keywords = item.get("关键词", [])
            if isinstance(keywords, list):
                keywords_str = ", ".join(keywords[:5])
            else:
                keywords_str = str(keywords)

            ws.cell(row=row_idx, column=1, value=item.get("标题", "")).alignment = self.LEFT_ALIGN
            ws.cell(row=row_idx, column=2, value=item.get("是否属于商机", "")).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=3, value=item.get("推荐跟进", "")).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=4, value=item.get("跟进优先级", "")).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=5, value=item.get("置信度", "")).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=6, value=item.get("人工复核", "")).alignment = self.CENTER_ALIGN
            ws.cell(row=row_idx, column=7, value=biz_types_str).alignment = self.LEFT_ALIGN
            ws.cell(row=row_idx, column=8, value=keywords_str).alignment = self.LEFT_ALIGN

            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = self.THIN_BORDER

        # 调整列宽
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 30

    # ========================
    # 文本报告
    # ========================
    def generate_text_report(self) -> str:
        """生成文本格式报告"""
        lines = []
        lines.append("=" * 60)
        lines.append("商机扫描汇总周报")
        lines.append("=" * 60)
        lines.append(f"生成时间: {self.summary.get('生成时间', '')}")
        lines.append(f"总处理数量: {self.summary.get('总处理数量', 0)}")

        lines.append("\n--- 商机识别结果 ---")
        biz = self.summary.get("商机识别结果", {})
        lines.append(f"  是: {biz.get('是', 0)}")
        lines.append(f"  否: {biz.get('否', 0)}")
        lines.append(f"  不确定: {biz.get('不确定', 0)}")

        lines.append("\n--- 推荐跟进结果 ---")
        rec = self.summary.get("推荐跟进结果", {})
        lines.append(f"  建议跟进: {rec.get('建议跟进', 0)}")
        lines.append(f"  可观察: {rec.get('可观察', 0)}")
        lines.append(f"  不建议跟进: {rec.get('不建议跟进', 0)}")

        lines.append("\n--- 优先级分布 ---")
        pri = self.summary.get("优先级分布", {})
        lines.append(f"  高: {pri.get('高', 0)}")
        lines.append(f"  中: {pri.get('中', 0)}")
        lines.append(f"  低: {pri.get('低', 0)}")

        lines.append(f"\n--- 人工复核池: {self.summary.get('进入人工复核池', 0)} ---")

        if self.high_value_opportunities:
            lines.append("\n--- 高价值商机（建议跟进）---")
            for i, opp in enumerate(self.high_value_opportunities[:10], 1):
                lines.append(f"  {i}. [{opp.get('跟进优先级', '')}] {opp.get('标题', '')[:50]}...")

        lines.append("\n" + "=" * 60)
        return "\n".join(lines)


if __name__ == "__main__":
    # 测试
    gen = WeeklyReportGenerator()

    # 模拟数据
    gen.data = [
        {
            "标题": "测试项目1：信息化项目第三方审计服务采购公告",
            "是否属于商机": "是",
            "推荐跟进": "建议跟进",
            "跟进优先级": "高",
            "置信度": "高",
            "商机类型": ["第三方审计服务", "信息化项目审计"],
            "关键词": ["信息化项目审计", "第三方审计", "结算审计"],
            "人工复核": "否",
            "判断依据": "正文明确采购内容为委托第三方独立审计机构",
            "风险提示": "无",
        },
        {
            "标题": "测试项目2：系统开发与审计打包",
            "是否属于商机": "否",
            "推荐跟进": "不建议跟进",
            "跟进优先级": "低",
            "置信度": "高",
            "商机类型": ["无"],
            "关键词": ["系统开发", "审计配合"],
            "人工复核": "否",
            "判断依据": "正文核心采购内容为系统开发",
            "风险提示": "无",
        },
    ]

    gen.aggregate()
    print(gen.generate_text_report())
    gen.generate_excel("test_report.xlsx")
    print("\n测试周报已生成: test_report.xlsx")