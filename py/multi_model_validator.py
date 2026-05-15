"""
multi_model_validator.py
同时发送数据到多个 AI 模型，对比返回结果
专用于：招投标商机识别（第三方独立审计机构可承接的信息化/信息系统/数据类审计服务）
"""

import os
import sys
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
import anthropic
import openai
import google.generativeai as genai
from cross_validator import cross_validate, print_report

# ========================
# 配置区
# ========================
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY   = os.getenv("OPENAI_API_KEY", "")
GOOGLE_API_KEY   = os.getenv("GOOGLE_API_KEY", "")

MODEL_CONFIG = {
    "claude-sonnet-4-6": {
        "provider": "anthropic",
        "model": "claude-sonnet-4-6-20250514",
        "max_tokens": 2048,
    },
    "gpt-4o": {
        "provider": "openai",
        "model": "gpt-4o",
        "max_tokens": 2048,
    },
    "gemini-2.0-flash": {
        "provider": "google",
        "model": "gemini-2.0-flash",
        "generation_config": {"max_output_tokens": 2048},
    },
}

# ========================
# 读取 prompt 文件
# ========================
PROMPT_FILE = r"C:\Users\Lenovo\Desktop\审计库\最终版.txt"

def load_system_prompt(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        return f.read()

SYSTEM_PROMPT = load_system_prompt(PROMPT_FILE)

# ========================
# Excel 读取
# ========================
def read_excel_data(filepath):
    """
    读取 Excel，每行结构：
    { "标题": ..., "正文": ..., "项目编号": ..., "日期": ... }
    第一行必须包含"标题"和"正文"列
    """
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    if "标题" not in headers or "正文" not in headers:
        raise ValueError("Excel 必须包含「标题」和「正文」两列")

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        data = dict(zip(headers, row))
        rows.append(data)
    return rows


# ========================
# 模型调用
# ========================
def call_anthropic(model, prompt_text, max_tokens=2048):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    message = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt_text}]
    )
    return message.content[0].text


def call_openai(model, prompt_text, max_tokens=2048):
    client = openai.OpenAI(api_key=OPENAI_API_KEY)
    response = client.chat.completions.create(
        model=model,
        max_tokens=max_tokens,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt_text}
        ]
    )
    return response.choices[0].message.content


def call_google(model, prompt_text):
    model_obj = genai.GenerativeModel(model)
    response = model_obj.generate_content(
        prompt_text,
        generation_config=genai.GenerationConfig(max_output_tokens=2048)
    )
    return response.text


def build_prompt(title, body):
    return f"标题：{title}\n\n正文：{body}"


def send_to_all_models(title, body):
    """并发发送给所有模型"""
    prompt = build_prompt(title, body)
    results = {}

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {
            executor.submit(_call_model, cfg, prompt): name
            for name, cfg in MODEL_CONFIG.items()
        }

        for future in as_completed(futures):
            name = futures[future]
            try:
                results[name] = future.result()
            except Exception as e:
                results[name] = f"[调用错误] {e}"

    return results


def _call_model(cfg, prompt):
    provider = cfg["provider"]
    model = cfg["model"]

    if provider == "anthropic":
        return call_anthropic(model, prompt, cfg.get("max_tokens", 2048))
    elif provider == "openai":
        return call_openai(model, prompt, cfg.get("max_tokens", 2048))
    elif provider == "google":
        return call_google(model, prompt)
    else:
        return f"[未知 provider] {provider}"


def print_result(title, results):
    print("\n" + "=" * 80)
    print(f"【标题】{title[:60]}{'...' if len(title) > 60 else ''}")
    print("=" * 80)
    for name, result in results.items():
        print(f"\n◆ {name}:")
        print(result)
        print("-" * 40)


def print_json_result(title, results):
    print("\n" + "=" * 80)
    print(f"【标题】{title[:60]}{'...' if len(title) > 60 else ''}")
    print("=" * 80)
    for name, result in results.items():
        print(f"\n◆ {name}:")
        print(result)


def save_results_to_file(results_by_title, cv_results_by_title, excel_path):
    """将结果追加写入 Excel（新建 sheet）"""
    wb = openpyxl.load_workbook(excel_path)
    # 创建新 sheet
    import datetime
    ts = datetime.datetime.now().strftime("%m%d_%H%M")
    sheet_name = f"验证结果_{ts}"
    ws = wb.create_sheet(sheet_name)

    # 表头
    headers = ["标题", "正文（摘要）", "最终判断", "最终推荐", "置信度", "是否人工复核"]
    for model_name in MODEL_CONFIG.keys():
        headers.append(f"{model_name}_结果")

    ws.append(headers)

    # 写数据
    for title, data in results_by_title.items():
        body_short = data["body"][:80] + "..." if len(data["body"]) > 80 else data["body"]
        cv = cv_results_by_title.get(title, {})

        row = [
            title,
            body_short,
            cv.get("final_judgment", ""),
            cv.get("final_recommendation", ""),
            cv.get("final_confidence", ""),
            "是" if cv.get("need_manual_review") else "否",
        ]
        for model_name in MODEL_CONFIG.keys():
            row.append(data["results"].get(model_name, ""))
        ws.append(row)

    wb.save(excel_path)
    print(f"\n结果已保存到: {excel_path} > {sheet_name}")


# ========================
# 主程序
# ========================
def main():
    print("=" * 60)
    print("  多模型验证工具 —— 招投标商机识别")
    print("=" * 60)

    # 检查 API Key
    missing = []
    if not ANTHROPIC_API_KEY: missing.append("ANTHROPIC_API_KEY")
    if not OPENAI_API_KEY: missing.append("OPENAI_API_KEY")
    if not GOOGLE_API_KEY: missing.append("GOOGLE_API_KEY")
    if missing:
        print(f"请先设置环境变量: {', '.join(missing)}")
        print("或直接在代码顶部填写 API Key")
        return

    # 加载 Prompt
    print(f"\n已加载 system prompt: {PROMPT_FILE}")

    # 选择模式
    print("\n请选择模式:")
    print("  [1] 读取 Excel 文件（批量验证）")
    print("  [2] 单条输入验证")
    mode = input("请输入（1/2）: ").strip()

    if mode == "1":
        # 批量模式
        excel_path = input("\n请输入 Excel 路径: ").strip().strip('"')
        if not excel_path:
            print("路径不能为空")
            return

        try:
            rows = read_excel_data(excel_path)
            print(f"已加载 {len(rows)} 条数据\n")
        except Exception as e:
            print(f"读取 Excel 失败: {e}")
            return

        # 确认
        start = input(f"即将验证全部 {len(rows)} 条，是否继续？（y/n）: ").strip().lower()
        if start != "y":
            print("已取消")
            return

        results_by_title = {}
        cv_results_by_title = {}
        for i, row in enumerate(rows):
            title = str(row.get("标题", "") or "")
            body  = str(row.get("正文", "") or "")
            if not title or not body:
                continue

            print(f"\n[{i+1}/{len(rows)}] 验证中: {title[:50]}...")
            results = send_to_all_models(title, body)
            results_by_title[title] = {"body": body, "results": results}

            # 交叉验证
            cv_result = cross_validate(title, results)
            cv_results_by_title[title] = cv_result
            print_report(cv_result)

        # 保存
        save_opt = input("\n是否将结果保存到 Excel？（y/n）: ").strip().lower()
        if save_opt == "y":
            save_results_to_file(results_by_title, cv_results_by_title, excel_path)

    elif mode == "2":
        # 单条模式
        print("\n请输入标题（回车换行，输入完成按 Ctrl+Z / Ctrl+D 结束）:")
        title = sys.stdin.read().strip()

        print("\n请输入正文内容:")
        body = sys.stdin.read().strip()

        if not title or not body:
            print("标题或正文不能为空")
            return

        print("\n正在发送给各模型...\n")
        results = send_to_all_models(title, body)
        print_json_result(title, results)

        # 交叉验证
        cv_result = cross_validate(title, results)
        print_report(cv_result)

    else:
        print("无效选项")

if __name__ == "__main__":
    # 依赖安装: pip install openpyxl anthropic openai google-generativeai
    main()